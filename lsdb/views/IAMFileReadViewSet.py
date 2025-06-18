from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework import viewsets
from azure.storage.blob import BlobServiceClient
from openpyxl import load_workbook
from io import BytesIO
import os

from lsdb.models.IAMFileRead import IAMFileRead
from lsdb.serializers import IAMFileReadSerializer

class IAMFileReadViewSet(viewsets.ModelViewSet):
    
    logging_methods = ['GET', 'POST', 'DELETE']
    queryset = IAMFileRead.objects.all()
    serializer_class = IAMFileReadSerializer


    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"status": "error", "msg": "No file uploaded."}, status=400)

        try:
            connect_str = 'DefaultEndpointsProtocol=https;AccountName=haveblueazdev;AccountKey=eP954sCH3j2+...==;EndpointSuffix=core.windows.net'
            container_name = 'testmedia1'

            blob_service_client = BlobServiceClient.from_connection_string(connect_str)
            blob_client = blob_service_client.get_blob_client(container=container_name, blob=file_obj.name)

            blob_client.upload_blob(file_obj, overwrite=True)

            return Response({"status": "success", "msg": f"{file_obj.name} uploaded to Azure Blob Storage."})
        except Exception as e:
            return Response({"status": "error", "msg": str(e)}, status=500)

    @action(detail=False, methods=['get'])
    def fileread(self, request):
        connect_str = 'DefaultEndpointsProtocol=https;AccountName=haveblueazdev;AccountKey=eP954sCH3j2+dbjzXxcAEj6n7vmImhsFvls+7ZU7F4THbQfNC0dULssGdbXdilTpMgaakIvEJv+QxCmz/G4Y+g==;EndpointSuffix=core.windows.net'
        container_name = 'testmedia1'
        try:
            blob_service_client = BlobServiceClient.from_connection_string(connect_str)
            container_client = blob_service_client.get_container_client(container_name)
            # Filter and get only Excel files
            excel_blobs = sorted(
                [blob for blob in container_client.list_blobs() if blob.name.endswith('.xlsx')],
                key=lambda x: x.last_modified,
                reverse=True
            )
            if not excel_blobs:
                return Response({"status": "error", "msg": "No Excel (.xlsx) files found."}, status=404)
            latest_blob = excel_blobs[0]
            blob_client = container_client.get_blob_client(latest_blob.name)
            file_data = blob_client.download_blob().readall()
            workbook = load_workbook(filename=BytesIO(file_data))
            sheet = workbook.active
            required_fields = ['Datetime', 'LED0', 'LED1', 'LED2', 'LED3', 'LED4', 'LED5', 'LED6', 'LED7']
            header = [cell.value for cell in sheet[1]]
            column_map = {name: idx for idx, name in enumerate(header) if name in required_fields}
            missing_fields = [field for field in required_fields if field not in column_map]
            if missing_fields:
                return Response({"status": "error", "msg": f"Missing fields: {missing_fields}"}, status=400)
            objs = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                objs.append(IAMFileRead(
                    datetime=row[column_map['Datetime']],
                    led0=row[column_map['LED0']],
                    led1=row[column_map['LED1']],
                    led2=row[column_map['LED2']],
                    led3=row[column_map['LED3']],
                    led4=row[column_map['LED4']],
                    led5=row[column_map['LED5']],
                    led6=row[column_map['LED6']],
                    led7=row[column_map['LED7']],
                ))
            IAMFileRead.objects.bulk_create(objs)
            return Response({
                "status": "success",
                "Processed Excel File": latest_blob.name,
                "Total Records Inserted": len(objs)
            })
        except Exception as e:
            return Response({"status": "error", "msg": f"Failed to process Excel file: {str(e)}"}, status=500)
