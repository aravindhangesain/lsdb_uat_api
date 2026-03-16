from rest_framework import viewsets
from lsdb.models import *
from lsdb.serializers import *
from django.utils import timezone
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from azure.storage.blob import BlobServiceClient
import os
from zipfile import ZipFile
import openpyxl
from django.db import transaction

class XlfilereadViewSet(viewsets.ModelViewSet):
    permission_classes = [AllowAny]
    logging_methods = ['GET','POST','DELETE']
    queryset = Xlfileread.objects.all()
    serializer_class = XlfilereadSerializer

    @action(detail=False, methods=['get'])
    def fileread(self, request):
        extract_path = "C:\\Users\\lsdb_local\\lsdb_uat_api\\lsdb\\xltest"
        os.makedirs(extract_path, exist_ok=True)

        # Azure connection string 
        connect_str = 'DefaultEndpointsProtocol=https;AccountName=haveblueazdev;AccountKey=eP954sCH3j2+dbjzXxcAEj6n7vmImhsFvls+7ZU7F4THbQfNC0dULssGdbXdilTpMgaakIvEJv+QxCmz/G4Y+g==;EndpointSuffix=core.windows.net'
        container_name = 'testmedia1'

        # Download from azure
        try:
            blob_service_client = BlobServiceClient.from_connection_string(connect_str)
            container_client = blob_service_client.get_container_client(container_name)

            zip_blobs = [
                blob for blob in container_client.list_blobs()
                if blob.name.endswith('.zip')
            ]
            if not zip_blobs:
                return Response({"status": "error", "msg": "No ZIP files."}, status=404)

            latest_blob = max(zip_blobs, key=lambda blob: blob.last_modified)
            blob_name = latest_blob.name

            # Downloading 
            zip_path = os.path.join(extract_path, os.path.basename(blob_name))
            blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

            with open(zip_path, "wb") as download_file:
                download_file.write(blob_client.download_blob().readall())
        except Exception as e:
            return Response({"status": "error", "msg": f"Azure download failed: {str(e)}"}, status=500)

        # Zip Extract
        try:
            with ZipFile(zip_path, 'r') as zobject:
                zobject.extractall(path=extract_path)
                extracted_files = zobject.namelist()
        except Exception as e:
            return Response({"status": "error", "msg": f"ZIP extracting failed: {str(e)}"}, status=500)

        excel_files = [os.path.join(extract_path, file) for file in extracted_files if file.endswith('.xlsx')]
        if not excel_files:
            return Response({"status": "error", "msg": "No Excel file."}, status=400)

        required_fields = ['serialnumber', 'projectnumber', 'customername', 'workorder', 'location']
        count = 0

        # read the excel
        for excel_file in excel_files:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            header = [cell.value for cell in sheet[1]]
            column_map = {name.lower(): index for index, name in enumerate(header) if name}

            missing_fields = [field for field in required_fields if field not in column_map]
            if missing_fields:
                continue

            objs = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                obj = Xlfileread(
                    serialnumber=row[column_map['serialnumber']],
                    projectnumber=row[column_map['projectnumber']],
                    customername=row[column_map['customername']],
                    workorder=row[column_map['workorder']],
                    location=row[column_map['location']],
                )
                objs.append(obj)

            Xlfileread.objects.bulk_create(objs)
            count += len(objs)

        return Response({
            "status": "success","Zip File Name": blob_name, "No.of.Records": count,"No.of.Excel": len(excel_files)
        })
        
    @transaction.atomic
    @action(detail=False, methods=['post','get'])
    def upload_excel(self, request):

        file = request.FILES.get('file')
        disposition = Disposition.objects.get(id=16)
        location = Location.objects.get(id=5)


        if not file:
            return Response({"status": "error", "msg": "Excel file required"}, status=400)

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
        except Exception as e:
            return Response({"status": "error", "msg": f"Invalid Excel file: {str(e)}"}, status=400)

        required_fields = ['name', 'description']

        header = [cell.value for cell in sheet[1]]
        column_map = {name.lower(): index for index, name in enumerate(header) if name}

        missing_fields = [field for field in required_fields if field not in column_map]
        if missing_fields:
            return Response({"status": "error", "msg": f"Missing columns: {missing_fields}"})

        objs = []
        count = 0


        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue

            objs.append(
                Asset(
                    name=row[column_map['name']],
                    description=row[column_map['description']] if 'description' in column_map else None,
                    last_action_datetime=timezone.now(),
                    disposition=disposition,
                    location=location,
                )
            )

        Asset.objects.bulk_create(objs,ignore_conflicts=True)
        count = len(objs)

        return Response({
            "status": "success",
            "records_inserted": count
        })
        
    @transaction.atomic    
    @action(detail=False, methods=['post','get'])
    def upload_calibration(self, request):

        file = request.FILES.get('file')
        disposition = Disposition.objects.get(id=16)
        location = Location.objects.get(id=5)


        if not file:
            return Response({"status": "error", "msg": "Excel file required"}, status=400)

        try:
            workbook = openpyxl.load_workbook(file,data_only=True)
            sheet = workbook.active
        except Exception as e:
            return Response({"status": "error", "msg": f"Invalid Excel file: {str(e)}"}, status=400)

        required_fields = ['asset_id','asset_number','asset_name', 'description','last_calibrated_date','schedule_for_calibration','manufacturer','serial_number']

        header = [cell.value for cell in sheet[1]]
        column_map = {name.lower(): index for index, name in enumerate(header) if name}

        missing_fields = [field for field in required_fields if field not in column_map]
        if missing_fields:
            return Response({"status": "error", "msg": f"Missing columns: {missing_fields}"})

        objs = []
        count = 0


        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            
            schedule = row[column_map['schedule_for_calibration']]
            schedule = None if schedule in [None, '', 'NA', 'N/A'] else int(schedule)

            objs.append(
                AssetCalibration(
                    asset_id=row[column_map['asset_id']],
                    asset_number=row[column_map['asset_number']],
                    asset_name=row[column_map['asset_name']],
                    description=row[column_map['description']] if 'description' in column_map else None,
                    last_action_datetime=timezone.now(),
                    location=location,
                    manufacturer=row[column_map['manufacturer']],
                    usage = None,
                    model = None,
                    serial_number=row[column_map['serial_number']],
                    is_calibration_required = True,
                    last_calibrated_date=row[column_map['last_calibrated_date']] or None,
                    schedule_for_calibration=schedule,
                    external_asset_required= None,
                    asset_type_id = None,
                    azurefile_id = None,
                    is_main_asset = True,
                    is_sub_asset = False,
                    is_rack = False,
                    disposition=disposition,
                   
                )
            )

        AssetCalibration.objects.bulk_create(objs,ignore_conflicts=True)
        count = len(objs)

        return Response({
            "status": "success",
            "records_inserted": count
        })
