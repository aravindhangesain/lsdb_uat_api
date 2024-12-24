from rest_framework import viewsets
from lsdb.models import StepResult_pichina, Unit_pichina
from lsdb.permissions import ConfiguredPermission
from lsdb.serializers import Unit_pichinaSerializer
from lsdb.serializers import UnitGroupedTraveler_pichinaSerializer
from lsdb.serializers import Note_pichinaSerializer
from lsdb.utils import ExcelFile
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from django_filters import rest_framework as filters
from django.shortcuts import get_object_or_404
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, NamedStyle, Side, PatternFill, Font, Alignment

class Unit_pichinaFilter(filters.FilterSet):
    # projects =

    class Meta:
        model = Unit_pichina
        fields = {
            'serial_number': ['exact', 'icontains'],
            # 'project_set',
        }


class Unit_pichinaViewSet(viewsets.ModelViewSet):
    logging_methods = ['POST', 'PUT', 'PATCH', 'DELETE']
    queryset = Unit_pichina.objects.all()
    serializer_class = Unit_pichinaSerializer
    filter_backends = (filters.DjangoFilterBackend,)
    filterset_class = Unit_pichinaFilter
    permission_classes = [ConfiguredPermission]
    

    @action(detail=True, methods=['get'],
            serializer_class=UnitGroupedTraveler_pichinaSerializer,
            )
    def grouped_history(self, request, pk=None):
        self.context = {'request': request}
        file = request.query_params.get('file', 'dummy')
        # queryset = Unit.objects.get(id=pk)#.prefetch_related('step_results__measurement_results')
        queryset = get_object_or_404(Unit_pichina, id=pk)
        serializer = self.serializer_class(queryset, many=False, context=self.context)
        if file.upper() == 'EXCEL':
            # wb = Workbook()
            myFile = ExcelFile()
            wb = myFile.workbook

            # Styles:
            hairline = Side(border_style="hair", color="000000")
            thin = Side(border_style="thin", color="000000")
            medium = Side(border_style="medium", color="000000")
            thick = Side(border_style='thick', color="000000")

            center_b = NamedStyle(name='center_b')
            center_b.border = Border(start=medium, end=medium, outline=True)
            center_b.fill = PatternFill("solid", fgColor="00FFFFFF")
            center_b.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=False, shrink_to_fit=True)
            center_b.font = Font(name='Arial', b=True, size=10)

            center_bl = NamedStyle(name='center_bl')
            center_bl.border = Border(start=medium, end=medium, outline=True, top=medium, bottom=medium)
            center_bl.alignment = Alignment(horizontal="center", vertical="center")
            center_bl.font = Font(name='Arial', b=True, size=22)

            center_bxl = NamedStyle(name='center_bxl')
            center_bxl.border = Border(start=medium, end=medium, outline=True, top=medium, bottom=medium)
            center_bxl.alignment = Alignment(horizontal="center", vertical="center",
                                             wrap_text=False,
                                             shrink_to_fit=True)
            center_bxl.font = Font(name='Arial', b=True, size=26)

            center = NamedStyle(name='center')
            center.border = Border(start=medium, end=medium, top=hairline, bottom=hairline, outline=True)
            center.alignment = Alignment(horizontal="center", vertical="center",
                                         wrap_text=False, shrink_to_fit=True)
            center.font = Font(name='Arial', b=False, size=10)

            leg_num = NamedStyle(name='leg_num')
            leg_num.alignment = Alignment(horizontal="center", vertical="center")
            # leg_num.border = Border(top=medium, left=medium, right=medium, bottom=medium)
            leg_num.border = Border(top=medium, left=medium, right=medium, bottom=medium, outline=True)
            leg_num.fill = PatternFill("solid", fgColor="00FFFF00")
            leg_num.font = Font(name='Arial', b=True, size=24, color="00000000")

            leg_head = NamedStyle(name='leg_head')
            leg_head.alignment = Alignment(horizontal="center", vertical="center")
            leg_head.border = Border(top=medium, left=medium, right=medium, bottom=medium, outline=True)
            leg_head.fill = PatternFill("solid", fgColor="00FFFF00")
            leg_head.font = Font(name='Arial', b=True, size=12, color="00000000")

            test_title = NamedStyle(name='test_title')
            test_title.border = Border(left=medium, right=medium, top=hairline, bottom=hairline, outline=True)
            test_title.alignment = Alignment(horizontal="left", vertical="center",
                                             wrap_text=False,
                                             shrink_to_fit=True)
            test_title.font = Font(name='Arial', b=False, size=11)

            tiny_grey = NamedStyle(name='tiny_grey')
            tiny_grey.border = Border(left=hairline, right=hairline, outline=True)
            tiny_grey.fill = PatternFill("solid", fgColor="00DCE6F2")
            tiny_grey.alignment = Alignment(horizontal="center", vertical="center",
                                            wrap_text=False,
                                            shrink_to_fit=True)
            tiny_grey.font = Font(name='Arial', b=False, size=8)

            test_data = NamedStyle(name='test_data')
            test_data.border = Border(left=hairline, right=hairline, top=hairline, bottom=hairline, outline=True)
            test_data.fill = PatternFill("solid", fgColor="00FFFFFF")
            test_data.alignment = Alignment(horizontal="center", vertical="center",
                                            wrap_text=False,
                                            shrink_to_fit=True)
            test_data.font = Font(name='Arial', b=False, size=8)

            wb.add_named_style(center_b)
            wb.add_named_style(center_bl)
            wb.add_named_style(center_bxl)
            wb.add_named_style(center)
            wb.add_named_style(leg_num)
            wb.add_named_style(leg_head)
            wb.add_named_style(test_title)
            wb.add_named_style(tiny_grey)
            wb.add_named_style(test_data)

            sheet = wb.active
            sheet.page_setup.fitToWidth = 1

            unit_type = serializer.data.pop('unit_type')
            calibration_results = serializer.data.pop('calibration_results')
            sequences_results = serializer.data.pop('sequences_results')
            module_property = unit_type.pop('module_property')
            sheet.merge_cells('A1:D2')
            sheet.merge_cells('E1:Q2')
            for i in range(1, 28):
                sheet.column_dimensions[get_column_letter(i)].width = 3.5
            if module_property.get('bifacial'):
                sheet['A1'] = 'Bifacial'
                sheet['A1'].style = 'center_bl'
            sheet['E1'] = serializer.data.pop('project_number')
            sheet['E1'].style = 'center_bl'

            for row in range(3, 9):
                for col in [[1, 4], [5, 8], [9, 12], [13, 17]]:
                    sheet.merge_cells(start_column=col[0], end_column=col[1], start_row=row, end_row=row)

            sheet.cell(row=3, column=1, value='Start Date').style = 'center_b'
            sheet.cell(row=3, column=5, value=serializer.data.pop('start_datetime')).style = 'center'
            sheet.cell(row=3, column=9, value='Mpp').style = 'center_b'
            sheet.cell(row=3, column=13, value=module_property.get('nameplate_pmax')).style = 'center'

            sheet.cell(row=4, column=1, value='Manufacturer').style = 'center_b'
            sheet.cell(row=4, column=5, value=unit_type.get('manufacturer_name')).style = 'center'
            sheet.cell(row=4, column=9, value='Voc').style = 'center_b'
            sheet.cell(row=4, column=13, value=module_property.get('voc')).style = 'center'

            sheet.cell(row=5, column=1, value='Module Type').style = 'center_b'
            sheet.cell(row=5, column=5, value=module_property.get('module_technology_name')).style = 'center'
            sheet.cell(row=5, column=9, value='Isc').style = 'center_b'
            sheet.cell(row=5, column=13, value=module_property.get('isc')).style = 'center'

            sheet.cell(row=6, column=1, value='Project Manager').style = 'center_b'
            sheet.cell(row=6, column=5, value=serializer.data.pop('project_manager')).style = 'center'
            sheet.cell(row=6, column=9, value='Vmp').style = 'center_b'
            sheet.cell(row=6, column=13, value=module_property.get('vmp')).style = 'center'

            sheet.cell(row=7, column=1, value='Calibration').style = 'center_b'
            sheet.cell(row=7, column=5, value='?????').style = 'center'
            sheet.cell(row=7, column=9, value='Imp').style = 'center_b'
            sheet.cell(row=7, column=13, value=module_property.get('imp')).style = 'center'

            sheet.cell(row=8, column=1, value='Dimensions').style = 'center_b'
            sheet.cell(row=8, column=5, value='{:n}x{:n}mm'.format(module_property.get('module_height'),
                                                                   module_property.get(
                                                                       'module_width'))).style = 'center'
            sheet.cell(row=8, column=9, value='Voltage').style = 'center_b'
            sheet.cell(row=8, column=13, value='{}V'.format(module_property.get('system_voltage'))).style = 'center'

            sheet.merge_cells(start_column=1, end_column=12, start_row=9, end_row=10)
            sheet.merge_cells(start_column=13, end_column=24, start_row=9, end_row=10)
            sheet.cell(row=9, column=1, value=serializer.data.pop('customer_name')).style = 'center_bxl'
            sheet.cell(row=9, column=13, value=serializer.data.pop('work_order_name')).style = 'center_bxl'

            sheet.merge_cells('R1:X2')
            sheet['R1'] = serializer.data.pop('serial_number')
            sheet['R1'].style = 'center_bxl'

            sheet.merge_cells('R3:X8')
            sheet['R3'] = serializer.data.pop('test_sequence_definition_name')
            sheet['R3'].style = 'center_bxl'
            sheet['R3'].alignment = Alignment(horizontal="center", vertical="center",
                                              wrap_text=True)
            sheet['R3'].fill = PatternFill("solid", fgColor="00003300")
            sheet['R3'].font = Font(name='Arial', b=True, size=26, color="00FFFFFF")

            # Apply heading styles:
            for row in range(3, 9):
                for col in ['A', 'L']:
                    cell = '{}{}'.format(col, row)
                    sheet[cell].style = 'center_b'
                for col in ['F', 'P']:
                    cell = '{}{}'.format(col, row)
                    sheet[cell].style = 'center'

            current_row = 11
            for sequences in sequences_results:
                sheet.merge_cells('A{}:B{}'.format(current_row, current_row + 1))
                sheet.merge_cells('C{}:X{}'.format(current_row, current_row))
                sheet['A{}'.format(current_row)] = sequences.get('linear_execution_group')
                sheet['A{}'.format(current_row)].style = 'leg_num'
                # leg_num = sequences.get('linear_execution_group')
                sheet['C{}'.format(current_row)] = sequences.get('name')
                sheet['C{}'.format(current_row)].style = 'leg_head'
                current_row += 1
                for col in range(0, 5):
                    sheet.merge_cells(start_column=5 + (col * 2), end_column=6 + (col * 2), start_row=current_row,
                                      end_row=current_row)

                sheet.cell(row=current_row, column=5, value='Date').style = 'tiny_grey'
                sheet.cell(row=current_row, column=7, value='Initials').style = 'tiny_grey'
                sheet.cell(row=current_row, column=9, value='Review Date').style = 'tiny_grey'
                sheet.cell(row=current_row, column=11, value='Reviewer').style = 'tiny_grey'
                sheet.cell(row=current_row, column=13, value='Result').style = 'tiny_grey'

                # leg_head.style = 'leg_head'
                # leg_head = sequences.get('name')
                start_row = current_row
                current_row += 1
                for result in sequences.get('procedure_results'):
                    sheet.merge_cells('A{}:D{}'.format(current_row, current_row))
                    sheet['A{}'.format(current_row)].style = 'test_title'
                    sheet['A{}'.format(current_row)] = result.get('procedure_definition_name')
                    for col in range(0, 5):
                        sheet.merge_cells(start_column=5 + (col * 2), end_column=6 + (col * 2), start_row=current_row,
                                          end_row=current_row)

                    sheet.cell(row=current_row, column=5, value=result.get('completion_date')).style = 'test_data'
                    sheet.cell(row=current_row, column=7, value=result.get('username')).style = 'test_data'
                    sheet.cell(row=current_row, column=9, value=result.get('review_datetime')).style = 'test_data'
                    sheet.cell(row=current_row, column=11, value=result.get('reviewed_by_user')).style = 'test_data'
                    sheet.cell(row=current_row, column=13, value=result.get('disposition_name')).style = 'test_data'

                    current_row += 1
                    if result.get('visualizer') == 'stress':
                        # need to add metadata for stressor rows here
                        steps = StepResult_pichina.objects.filter(procedure_result__id=result.get('id')).order_by(
                            'linear_execution_group')
                        # steps = steps.exclude(name__iexact='test end')
                        # steps = steps.exclude(name__iexact='test start')
                        for step in steps:
                            sheet.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=4)
                            sheet.cell(row=current_row, column=1, value=step.name).style = 'test_title'
                            for col in range(0, 5):
                                sheet.merge_cells(start_column=5 + (col * 2), end_column=6 + (col * 2),
                                                  start_row=current_row, end_row=current_row)
                            current_row += 1
                sheet.merge_cells(start_column=15, end_column=24, start_row=start_row, end_row=current_row - 1)
                sheet.cell(row=start_row, column=15, value=' ').style = 'center_b'

            # mem_file = BytesIO(save_virtual_workbook(wb))
            filename = timezone.now().strftime('%b-%d-%Y-%H%M%S')
            # response = HttpResponse(mem_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            # response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(filename)
            # return response
            return myFile.get_response(filename)
        else:
            return Response(serializer.data)
        

    @action(detail=True, methods=['get', ],
            serializer_class=Note_pichinaSerializer,
            )
    def notes(self, request, pk=None):
        unit = Unit_pichina.objects.get(id=pk)
        self.context = {'request': request}
        serializer = Note_pichinaSerializer(unit.notes.all(), many=True, context=self.context)
        return Response(serializer.data)


