from django.shortcuts import get_object_or_404
from rest_framework import viewsets
from lsdb.models import *
from lsdb.serializers import *
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.utils import timezone
import zipfile
from io import BytesIO
from openpyxl import Workbook
from PIL import Image, ImageOps
from django.db.models import F, Value,Q
from django.db.models.functions import Replace, Lower
from openpyxl import Workbook
from io import BytesIO
from django.http import HttpResponse
# from zipstream import ZipFile
from django.http import StreamingHttpResponse
from lsdb.utils.ZipFileUtils import workorder_download_file



class ProjectdownloadViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    serializer_class = ProjectdownloadSerializer
    lookup_field = 'number'

    # def retrieve(self, request, number=None):
    #     project = self.get_object()
    #     data = {
    #         "project_id": project.id,
    #         "project_number": project.number,
    #         "workorders": []
    #     }
    #     for workorder in project.workorder_set.all():
    #         procedures = ProcedureResult.objects.filter(
    #             work_order=workorder.id
    #         ).values(
    #             'procedure_definition__id',
    #             'procedure_definition__name',
    #             'name'
    #         ).exclude(group_id=45)
    #         procedures_dict = {}
    #         for proc in procedures:
    #             pid = proc['procedure_definition__id']
    #             pname = proc.get('name')
    #             if pid not in procedures_dict:
    #                 procedures_dict[pid] = {
    #                     "procedure_definition_id": pid,
    #                     "procedure_definition_name": proc['procedure_definition__name'],
    #                     "procedure_names": []
    #                 }
    #             if pname and pname not in procedures_dict[pid]["procedure_names"]:
    #                 procedures_dict[pid]["procedure_names"].append(pname)
    #         serial_numbers = list(
    #             Unit.objects.filter(procedureresult__work_order=workorder.id).values_list('serial_number', flat=True).distinct())
    #         workorder_data = {
    #             "workorder_id": workorder.id,
    #             "workorder_name": workorder.name,
    #             "serial_numbers": serial_numbers,
    #             "procedure_definitions": list(procedures_dict.values())
    #         }
    #         data["workorders"].append(workorder_data)
    #     return Response(data)

    def retrieve(self, request, number=None):
        project = self.get_object()
        data = {
            "project_id": project.id,
            "project_number": project.number,
            "workorders": []
        }
        for workorder in project.workorder_set.all():
            serial_numbers = Unit.objects.filter(
                procedureresult__work_order=workorder.id
            ).values_list('serial_number', flat=True).distinct()
            details_list = []
            for serial in serial_numbers:
                procedure_names = ProcedureResult.objects.filter(
                    work_order=workorder.id,
                    unit__serial_number=serial
                ).exclude(group_id=45).values_list('name', flat=True).distinct()
                procedures_list = []
                for proc_name in procedure_names:
                    proc_defs = ProcedureResult.objects.filter(
                        work_order=workorder.id,
                        unit__serial_number=serial,
                        name=proc_name
                    ).exclude(group_id=45).values(
                        'procedure_definition__id',
                        'procedure_definition__name'
                    ).order_by('procedure_definition__id').distinct('procedure_definition__id')
                    procedure_def_details = [
                        {
                            "procedure_definition_id": p['procedure_definition__id'],
                            "procedure_definition_name": p['procedure_definition__name']
                        }
                        for p in proc_defs
                    ]
                    procedures_list.append({
                        "procedure_name": proc_name,
                        "procedure_definition_details": procedure_def_details
                    })
                details_list.append({
                    "serial_numbers": serial,
                    "procedures": procedures_list
                })
            workorder_data = {
                "workorder_id": workorder.id,
                "workorder_name": workorder.name,
                "details": details_list
            }
            data["workorders"].append(workorder_data)
        return Response(data)

    @action(detail=True, methods=['post'], url_path='download')
    def download(self, request, number=None):
        # ----------------------------
        # Helpers
        # ----------------------------
        def normalize_name(queryset):
            return queryset.annotate(
                normalized_name=Lower(
                    Replace(
                        Replace(F("name"), Value("-"), Value("")),
                        Value(" "), Value("")
                    )
                )
            )

        def normalize_value(val: str) -> str:
            return val.replace("-", "").replace(" ", "").lower()

        def safe_excel_value(val):
            if hasattr(val, "all"):
                return ", ".join(str(v) for v in val.all())
            if isinstance(val, (str, int, float, type(None))):
                return val
            return str(val)

        # ----------------------------
        # Validate payload
        # ----------------------------
        workorder_id = request.data.get("workorder_id")
        serial_numbers = request.data.get("serial_numbers", [])
        procedure_names = request.data.get("procedure_names", [])
        procedure_definition_ids = request.data.get("procedure_definition_id", [])
        adjust_images = False

        allowed_proc_def = [2, 3, 14, 54, 50, 62, 33, 49, 21, 38, 48, 12, 18, 37]
        project = get_object_or_404(Project, number=number)
        work_order = get_object_or_404(project.workorder_set, id=workorder_id)

        if workorder_id is not None and not serial_numbers:

            work_orders = WorkOrder.objects.filter(id=workorder_id)
            return workorder_download_file(work_orders)

            # procedures=ProcedureResult.objects.filter(
            #             Q(work_order_id=workorder_id) & Q(procedure_definition__id__in=allowed_proc_def)).order_by(
            #             'test_sequence_definition', 'linear_execution_group').select_related(
            #             'unit').prefetch_related('stepresult_set__measurementresult_set').distinct().values_list('id',flat=True)
            
        else:

                project = get_object_or_404(Project, number=number)
                work_order = get_object_or_404(project.workorder_set, id=workorder_id)

            

                # ----------------------------
                # Build procedures grouped by procedure_definition_id
                # ----------------------------
                procedures = []
                if not procedure_definition_ids:
                    procedure_results = ProcedureResult.objects.filter(
                        work_order=work_order, procedure_definition_id__in=allowed_proc_def
                    ).exclude(group_id=45)
                    proc_defs = procedure_results.values_list("procedure_definition_id", flat=True).distinct()
                    for proc_def_id in proc_defs:
                        proc_names = (
                            procedure_results.filter(procedure_definition_id=proc_def_id)
                            .values_list("name", flat=True)
                            .distinct()
                        )
                        procedures.append({
                            "procedure_definition_id": proc_def_id,
                            "procedure_names": list(proc_names)
                        })
                else:
                    for proc_def_id in procedure_definition_ids:
                        procedures.append({
                            "procedure_definition_id": proc_def_id,
                            "procedure_names": procedure_names
                        })

                files_to_return = []
                extra_files_exist = False

        # ----------------------------
                # Loop over each procedure_definition_id
                # ----------------------------
                for proc in procedures:
                    
                    
                    # ----------------------------
                    # Determine units
                    # ----------------------------
                    if serial_numbers:
                        procedure_def_id = proc.get("procedure_definition_id")
                        procedure_names_list = proc.get("procedure_names", [])
                        procedure_def = get_object_or_404(ProcedureDefinition, id=procedure_def_id)


                        units = Unit.objects.filter(serial_number__in=serial_numbers).distinct()
                    else:
                        procedure_obj = get_object_or_404(ProcedureResult, id=proc)
                        procedure_def_id = procedure_obj.procedure_definition_id
                        procedure_names_list = procedure_obj.name
                        procedure_def = get_object_or_404(ProcedureDefinition, id=procedure_def_id)

                        units = Unit.objects.filter(procedureresult__id__in=procedures)



                    # ----------------------------
                    # Create normalized procedure names list for filtering
                    # ----------------------------
                    normalized_proc_names = [normalize_value(pn) for pn in procedure_names_list] if procedure_names_list else [None]

                    # ----------------------------
                    # Excel sheet creation (existing logic remains unchanged)
                    # ----------------------------
                    excel_file = Workbook()
                    sheet = excel_file.active

                    # ========================================================
                    # I-V Flash Procedure
                    # ========================================================
                    if "I-V" in procedure_def.name:
                        sheet.title = "Flash"
                        sheet.append([
                            "Serial Number", "TSD", "LEG", "final_result",
                            "Pmp", "Voc", "Vmp", "Isc", "Imp", "Irradiance", "Temperature"
                        ])
                        for unit in units:
                            tests = normalize_name(
                                ProcedureResult.objects.filter(
                                    unit=unit,
                                    procedure_definition_id=procedure_def.id,
                                )
                            )
                            if normalized_proc_names != [None]:
                                tests = tests.filter(normalized_name__in=normalized_proc_names)
                            tests = tests.order_by("test_sequence_definition", "linear_execution_group")

                            for test in tests:
                                final_result_value = ProcedureResult_FinalResult.objects.filter(
                                    procedure_result_id=test.id
                                ).values_list("final_result", flat=True).first() or "N/A"

                                for step_result in test.stepresult_set.all().exclude(archived=True):
                                    data = [
                                        safe_excel_value(unit.serial_number),
                                        safe_excel_value(test.test_sequence_definition.name),
                                        safe_excel_value(test.name),
                                        safe_excel_value(final_result_value)
                                    ]

                                    for measurement in step_result.measurementresult_set.all().order_by("report_order"):
                                        if measurement.measurement_result_type.name == "result_files":
                                            if measurement.result_files.all().count():
                                                for azurefile in measurement.result_files.all():
                                                    path = "Flash Data/"
                                                    if "200" in step_result.name:
                                                        path += "200W/"
                                                    elif "Rear" in test.procedure_definition.name:
                                                        path += "Rear/"
                                                    path += "FLASH/{}/{}".format(
                                                        test.test_sequence_definition.name,
                                                        test.name
                                                    )
                                                    file_bytes = azurefile.file.file.read()
                                                    files_to_return.append({
                                                        "name": f"{work_order.project.number}/{work_order.name}/{path}/{azurefile.file.name}",
                                                        "content": file_bytes,
                                                        "force_zip": True
                                                    })
                                        elif measurement.measurement_result_type.name == "result_datetime":
                                            continue
                                        else:
                                            raw_value = getattr(
                                                measurement,
                                                measurement.measurement_result_type.name,
                                                "",
                                            )
                                            data.append(safe_excel_value(raw_value))
                                    sheet.append(data)

                    # ========================================================
                    # Visual Inspection
                    # ========================================================
                    elif "Visual Inspection" in procedure_def.name:
                        sheet.title = "Visual Inspection"
                        sheet.append([
                            "Serial Number", "TSD", "LEG", "final_result",
                            "Defect", "Category", "Notes", "Images"
                        ])
                        for unit in units:
                            tests = normalize_name(
                                ProcedureResult.objects.filter(
                                    unit=unit,
                                    procedure_definition_id=procedure_def.id,
                                )
                            )
                            if normalized_proc_names != [None]:
                                tests = tests.filter(normalized_name__in=normalized_proc_names)
                            tests = tests.order_by("test_sequence_definition", "linear_execution_group")

                            for test in tests:
                                final_result_value = ProcedureResult_FinalResult.objects.filter(
                                    procedure_result_id=test.id
                                ).values_list("final_result", flat=True).first() or "N/A"

                                skip = False
                                for step_result in test.stepresult_set.all().exclude(archived=True):
                                    data = [
                                        unit.serial_number,
                                        safe_excel_value(test.test_sequence_definition.name),
                                        safe_excel_value(test.name),
                                        safe_excel_value(final_result_value)
                                    ]

                                    for measurement in step_result.measurementresult_set.all().order_by("report_order"):
                                        if step_result.name == "Inspect Module":
                                            if getattr(measurement, measurement.measurement_result_type.name):
                                                skip = True
                                                data.append("No Defects Observed")
                                                sheet.append(data)
                                        elif skip:
                                            break
                                        elif measurement.measurement_result_type.name == "result_files":
                                            if measurement.result_files.all().count():
                                                for azurefile in measurement.result_files.all():
                                                    file_bytes = azurefile.file.file.read()
                                                    files_to_return.append({
                                                        "name": f"{work_order.project.number}/{work_order.name}/VI Images/{azurefile.file.name}",
                                                        "content": file_bytes,
                                                        "force_zip": True
                                                    })
                                                    data.append(azurefile.file.name)
                                        elif measurement.result_defect is not None and measurement.measurement_result_type.name == "result_defect":
                                            data.append(safe_excel_value(measurement.result_defect.short_name))
                                            data.append(safe_excel_value(measurement.result_defect.category))
                                        else:
                                            raw_value = getattr(
                                                measurement,
                                                measurement.measurement_result_type.name,
                                                "",
                                            )
                                            data.append(safe_excel_value(raw_value))

                                    if skip:
                                        break
                                    sheet.append(data)

                    # ========================================================
                    # Wet Leakage
                    # ========================================================
                    elif "Wet Leakage" in procedure_def.name:
                        sheet.title = "Wet Leakage"
                        sheet.append([
                            "Serial Number", "TSD", "LEG", "final_result",
                            "Insulation Resistance", "Passed?", "Test Voltage",
                            "Leakage Current", "Current Trip Setpoint", "Water Temperature"
                        ])
                        for unit in units:
                            tests = normalize_name(
                                ProcedureResult.objects.filter(
                                    unit=unit,
                                    procedure_definition_id=procedure_def.id,
                                )
                            )
                            if normalized_proc_names != [None]:
                                tests = tests.filter(normalized_name__in=normalized_proc_names)
                            tests = tests.order_by("test_sequence_definition", "linear_execution_group")

                            for test in tests:
                                final_result_value = ProcedureResult_FinalResult.objects.filter(
                                    procedure_result_id=test.id
                                ).values_list("final_result", flat=True).first() or "N/A"

                                for step_result in test.stepresult_set.all().exclude(archived=True):
                                    data = [
                                        safe_excel_value(unit.serial_number),
                                        safe_excel_value(test.test_sequence_definition.name),
                                        safe_excel_value(test.name),
                                        safe_excel_value(final_result_value)
                                    ]
                                    for measurement in step_result.measurementresult_set.all().order_by("report_order"):
                                        raw_value = getattr(
                                            measurement,
                                            measurement.measurement_result_type.name,
                                            "",
                                        )
                                        data.append(safe_excel_value(raw_value))
                                    sheet.append(data)

                    # ========================================================
                    # EL Image
                    # ========================================================
                    elif "EL Image" in procedure_def.name:
                        sheet.title = "EL Image"
                        sheet.append([
                            "Serial Number", "TSD", "LEG", "final_result",
                            "Measurement Values"
                        ])
                        for unit in units:
                            tests = normalize_name(
                                ProcedureResult.objects.filter(
                                    unit=unit,
                                    procedure_definition_id=procedure_def.id,
                                )
                            )
                            if normalized_proc_names != [None]:
                                tests = tests.filter(normalized_name__in=normalized_proc_names)
                            tests = tests.order_by("test_sequence_definition", "linear_execution_group")

                            for test in tests:
                                final_result_value = ProcedureResult_FinalResult.objects.filter(
                                    procedure_result_id=test.id
                                ).values_list("final_result", flat=True).first() or "N/A"

                                for step_result in test.stepresult_set.all().exclude(archived=True):
                                    data = [
                                        safe_excel_value(unit.serial_number),
                                        safe_excel_value(test.test_sequence_definition.name),
                                        safe_excel_value(test.name),
                                        safe_excel_value(final_result_value)
                                    ]

                                    for measurement in step_result.measurementresult_set.all().order_by("report_order"):
                                        if measurement.measurement_result_type.name == "result_files":
                                            if measurement.result_files.all().count():
                                                for azurefile in measurement.result_files.all():
                                                    if "RAW" in azurefile.file.name:
                                                        continue

                                                    filetype = "DataFiles" if azurefile.file.name.lower().endswith(("xls", "xlsx", "txt", "csv")) else "ImageFiles"

                                                    name = azurefile.file.name
                                                    if filetype == "ImageFiles":
                                                        temp = azurefile.file.name.split(".")
                                                        base_name = ".".join(temp[:-1])
                                                        name = "{}-{}-{}.{}".format(
                                                            base_name,
                                                            test.test_sequence_definition.name,
                                                            test.name,
                                                            temp[-1]
                                                        )

                                                    bytes_io = BytesIO(azurefile.file.file.read())
                                                    bytes_io.seek(0)

                                                    try:
                                                        temp_image = Image.open(bytes_io)
                                                        if adjust_images:
                                                            width, height = temp_image.size
                                                            temp_image = temp_image.rotate(-90, expand=True)
                                                            temp_image = temp_image.resize((height, width))
                                                            temp_image = ImageOps.grayscale(temp_image)
                                                            temp_image = ImageOps.autocontrast(temp_image)

                                                        image_bytes = BytesIO()
                                                        temp_image.save(image_bytes, format="jpeg", quality=75)
                                                        temp_image.close()
                                                        content = image_bytes.getvalue()
                                                        image_bytes.close()
                                                    except Exception:
                                                        bytes_io.seek(0)
                                                        content = bytes_io.getvalue()
                                                    bytes_io.close()

                                                    files_to_return.append({
                                                        "name": f"{work_order.project.number}/{work_order.name}/EL Images/{name}",
                                                        "content": content,
                                                        "force_zip": True
                                                    })
                                        else:
                                            raw_value = getattr(
                                                measurement,
                                                measurement.measurement_result_type.name,
                                                "",
                                            )
                                            data.append(safe_excel_value(raw_value))
                                    sheet.append(data)

                    else:
                        return Response(
                            {"error": f"Unsupported procedure type for {procedure_def.name}"},
                            status=400
                        )

                    # ----------------------------
                    # Save Excel file
                    # ----------------------------
                    file_stream = BytesIO()
                    excel_file.save(file_stream)
                    file_stream.seek(0)

                    file_name = f"{work_order.name}_{procedure_def.name}.xlsx"

                    files_to_return.append({
                        "name": file_name,
                        "content": file_stream.getvalue(),
                        "force_zip": False
                    })
                    file_stream.close()

                # ----------------------------
                # ✅ Remove duplicate Excel files
                # ----------------------------
                unique_files = []
                seen_names = set()

                for f in files_to_return:
                    if f["name"] not in seen_names:
                        unique_files.append(f)
                        seen_names.add(f["name"])

                files_to_return = unique_files

                # ----------------------------
                # Return single Excel or ZIP
                # ----------------------------
                if len(files_to_return) == 1 and not extra_files_exist:
                    file = files_to_return[0]
                    response = HttpResponse(
                        file["content"],
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    response["Content-Disposition"] = f'attachment; filename="{file["name"]}"'
                    return response
                
                mem_zip = BytesIO()
                with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                    for file in files_to_return:
                        zf.writestr(file["name"], file["content"])
                mem_zip.seek(0)

                response = HttpResponse(mem_zip.getvalue(), content_type="application/x-zip-compressed")
                response["Content-Disposition"] = (
                    f'attachment; filename="{work_order.project.number}-{timezone.now().strftime("%b-%d-%Y-%H%M%S")}.zip"'
                )
                return response
                
                # z = ZipFile(mode='w', compression=zipfile.ZIP_DEFLATED)
                # for file in files_to_return:
                #     z.write_iter(file["name"], BytesIO(file["content"]))
                # response = StreamingHttpResponse(z, content_type='application/zip')
                # response["Content-Disposition"] = (
                #     f'attachment; filename="{work_order.project.number}-{timezone.now().strftime("%b-%d-%Y-%H%M%S")}.zip"'
                # )
                # return response

       



    # @action(detail=True, methods=['post'], url_path='download')
    # def download(self, request, number=None):
    #     """
    #     Endpoint: POST /api/1.0/projectdownload/<pk>/download/

    #     Example Body JSON:
    #     {
    #         "workorder_id": 1,
    #         "procedures": [
    #             {
    #             "procedure_definition_id": 2,
    #             "procedure_names": ["Pre-Stress","Post-Stress"]
    #             },
    #             {
    #             "procedure_definition_id": 3,
    #             "procedure_names": ["Pre Light Soak","Post Light Soak"]
    #             }
    #         ]
    #     }
    #     """

    #     # ----------------------------
    #     # Helpers
    #     # ----------------------------
    #     def normalize_name(queryset):
    #         return queryset.annotate(
    #             normalized_name=Lower(
    #                 Replace(
    #                     Replace(F("name"), Value("-"), Value("")),
    #                     Value(" "), Value("")
    #                 )
    #             )
    #         )

    #     def normalize_value(val: str) -> str:
    #         return val.replace("-", "").replace(" ", "").lower()

    #     def safe_excel_value(val):
    #         if hasattr(val, "all"):
    #             return ", ".join(str(v) for v in val.all())
    #         if isinstance(val, (str, int, float, type(None))):
    #             return val
    #         return str(val)

    #     # ----------------------------
    #     # Validate body payload
    #     # ----------------------------
    #     workorder_id = request.data.get("workorder_id")
    #     procedures = request.data.get("procedures", [])
    #     adjust_images = False

    #     if not workorder_id:
    #         return Response({"error": "workorder id are required"}, status=400)

    #     project = get_object_or_404(Project, number=number)
    #     work_order = get_object_or_404(project.workorder_set, id=workorder_id)

    #     allowed_proc_def = [2, 3, 14, 54, 50, 62, 33, 49, 21, 38, 48, 12, 18, 37]

    #     if not procedures:
    #         procedures = []
    #         procedure_results = ProcedureResult.objects.filter(
    #             work_order=work_order, procedure_definition_id__in=allowed_proc_def
    #         ).exclude(group_id=45)
    #         proc_defs = procedure_results.values_list("procedure_definition_id", flat=True).distinct()

    #         for proc_def_id in proc_defs:
    #             proc_names = (
    #                 procedure_results.filter(procedure_definition_id=proc_def_id)
    #                 .values_list("name", flat=True)
    #                 .distinct()
    #             )
    #             procedures.append({
    #                 "procedure_definition_id": proc_def_id,
    #                 "procedure_names": list(proc_names)
    #             })

    #     files_to_return = []
    #     extra_files_exist = False

    #     # ----------------------------
    #     # Loop over procedures
    #     # ----------------------------
    #     for proc in procedures:
    #         procedure_def_id = proc.get("procedure_definition_id")
    #         procedure_names = proc.get("procedure_names", [])

    #         procedure_def = get_object_or_404(ProcedureDefinition, id=procedure_def_id)

    #         # ----------------------------
    #         # Determine if only workorder_id passed
    #         # ----------------------------
    #         if not request.data.get("procedures"):
    #             procedure_results = ProcedureResult.objects.filter(
    #                 work_order=work_order,
    #                 procedure_definition=procedure_def
    #             ).exclude(group_id=45)

    #             units = Unit.objects.filter(procedureresult__in=procedure_results).distinct()
    #             procedure_names_to_use = [None]

    #         else:
    #             procedure_names_to_use = procedure_names
    #             units = Unit.objects.filter(
    #                 procedureresult__work_order=work_order
    #             ).distinct()

    #         # ----------------------------
    #         # Loop over procedure_names if provided
    #         # ----------------------------
    #         for procedure_name in procedure_names_to_use:
    #             normalized_proc_name = normalize_value(procedure_name) if procedure_name else None

    #             if request.data.get("procedures"):
    #                 excel_file = Workbook()
    #                 sheet = excel_file.active
    #             else:
    #                 excel_file = Workbook()
    #                 sheet = excel_file.active

    #             # ========================================================
    #             # I-V Flash Procedure
    #             # ========================================================
    #             if "I-V" in procedure_def.name:
    #                 sheet.title = "Flash"
    #                 sheet.append([
    #                     "Serial Number", "TSD", "LEG", "final_result",
    #                     "Pmp", "Voc", "Vmp", "Isc", "Imp", "Irradiance", "Temperature"
    #                 ])

    #                 for unit in units:
    #                     tests = normalize_name(
    #                         ProcedureResult.objects.filter(
    #                             unit=unit,
    #                             procedure_definition_id=procedure_def.id,
    #                         )
    #                     )
    #                     if normalized_proc_name:
    #                         tests = tests.filter(normalized_name=normalized_proc_name)
    #                     tests = tests.order_by("test_sequence_definition", "linear_execution_group")

    #                     for test in tests:
    #                         final_result_value = ProcedureResult_FinalResult.objects.filter(
    #                             procedure_result_id=test.id
    #                         ).values_list("final_result", flat=True).first() or "N/A"

    #                         for step_result in test.stepresult_set.all().exclude(archived=True):
    #                             data = [
    #                                 safe_excel_value(unit.serial_number),
    #                                 safe_excel_value(test.test_sequence_definition.name),
    #                                 safe_excel_value(test.name),
    #                                 safe_excel_value(final_result_value)
    #                             ]

    #                             for measurement in step_result.measurementresult_set.all().order_by("report_order"):
    #                                 if measurement.measurement_result_type.name == "result_files":
    #                                     if measurement.result_files.all().count():
    #                                         for azurefile in measurement.result_files.all():
    #                                             path = "Flash Data/"
    #                                             if "200" in step_result.name:
    #                                                 path += "200W/"
    #                                             elif "Rear" in test.procedure_definition.name:
    #                                                 path += "Rear/"
    #                                             path += "FLASH/{}/{}".format(
    #                                                 test.test_sequence_definition.name,
    #                                                 test.name
    #                                             )
    #                                             file_bytes = azurefile.file.file.read()

    #                                             # collect into files_to_return instead of writing directly
    #                                             files_to_return.append({
    #                                                 "name": f"{work_order.project.number}/{work_order.name}/{path}/{azurefile.file.name}",
    #                                                 "content": file_bytes,
    #                                                 "force_zip": True
    #                                             })
    #                                 elif measurement.measurement_result_type.name == "result_datetime":
    #                                     continue
    #                                 else:
    #                                     raw_value = getattr(
    #                                         measurement,
    #                                         measurement.measurement_result_type.name,
    #                                         "",
    #                                     )
    #                                     data.append(safe_excel_value(raw_value))
    #                             sheet.append(data)

    #             # ========================================================
    #             # Visual Inspection
    #             # ========================================================
    #             elif "Visual Inspection" in procedure_def.name:
    #                 sheet.title = "Visual Inspection"
    #                 sheet.append([
    #                     "Serial Number", "TSD", "LEG", "final_result",
    #                     "Defect", "Category", "Notes", "Images"
    #                 ])
    #                 for unit in units:
    #                     tests = normalize_name(
    #                         ProcedureResult.objects.filter(
    #                             unit=unit,
    #                             procedure_definition_id=procedure_def.id,
    #                         )
    #                     )
    #                     if normalized_proc_name:
    #                         tests = tests.filter(normalized_name=normalized_proc_name)
    #                     tests = tests.order_by("test_sequence_definition", "linear_execution_group")

    #                     for test in tests:
    #                         final_result_value = ProcedureResult_FinalResult.objects.filter(
    #                             procedure_result_id=test.id
    #                         ).values_list("final_result", flat=True).first() or "N/A"

    #                         skip = False
    #                         for step_result in test.stepresult_set.all().exclude(archived=True):
    #                             data = [
    #                                 unit.serial_number,
    #                                 safe_excel_value(test.test_sequence_definition.name),
    #                                 safe_excel_value(test.name),
    #                                 safe_excel_value(final_result_value)
    #                             ]

    #                             for measurement in step_result.measurementresult_set.all().order_by("report_order"):
    #                                 if step_result.name == "Inspect Module":
    #                                     if getattr(measurement, measurement.measurement_result_type.name):
    #                                         skip = True
    #                                         data.append("No Defects Observed")
    #                                         sheet.append(data)

    #                                 elif skip:
    #                                     break

    #                                 elif measurement.measurement_result_type.name == "result_files":
    #                                     if measurement.result_files.all().count():
    #                                         for azurefile in measurement.result_files.all():
    #                                             file_bytes = azurefile.file.file.read()
    #                                             files_to_return.append({
    #                                                 "name": f"{work_order.project.number}/{work_order.name}/VI Images/{azurefile.file.name}",
    #                                                 "content": file_bytes,
    #                                                 "force_zip": True
    #                                             })
    #                                             data.append(azurefile.file.name)

    #                                 elif measurement.result_defect is not None and measurement.measurement_result_type.name == "result_defect":
    #                                     data.append(safe_excel_value(measurement.result_defect.short_name))
    #                                     data.append(safe_excel_value(measurement.result_defect.category))

    #                                 else:
    #                                     raw_value = getattr(
    #                                         measurement,
    #                                         measurement.measurement_result_type.name,
    #                                         "",
    #                                     )
    #                                     data.append(safe_excel_value(raw_value))

    #                             if skip:
    #                                 break
    #                             sheet.append(data)

    #             # ========================================================
    #             # Wet Leakage
    #             # ========================================================
    #             elif "Wet Leakage" in procedure_def.name:
    #                 sheet.title = "Wet Leakage"
    #                 sheet.append([
    #                     "Serial Number", "TSD", "LEG", "final_result",
    #                     "Insulation Resistance", "Passed?", "Test Voltage",
    #                     "Leakage Current", "Current Trip Setpoint", "Water Temperature"
    #                 ])
    #                 for unit in units:
    #                     tests = normalize_name(
    #                         ProcedureResult.objects.filter(
    #                             unit=unit,
    #                             procedure_definition_id=procedure_def.id,
    #                         )
    #                     )
    #                     if normalized_proc_name:
    #                         tests = tests.filter(normalized_name=normalized_proc_name)
    #                     tests = tests.order_by("test_sequence_definition", "linear_execution_group")

    #                     for test in tests:
    #                         final_result_value = ProcedureResult_FinalResult.objects.filter(
    #                             procedure_result_id=test.id
    #                         ).values_list("final_result", flat=True).first() or "N/A"

    #                         for step_result in test.stepresult_set.all().exclude(archived=True):
    #                             data = [
    #                                 safe_excel_value(unit.serial_number),
    #                                 safe_excel_value(test.test_sequence_definition.name),
    #                                 safe_excel_value(test.name),
    #                                 safe_excel_value(final_result_value)
    #                             ]

    #                             for measurement in step_result.measurementresult_set.all().order_by("report_order"):
    #                                 raw_value = getattr(
    #                                     measurement,
    #                                     measurement.measurement_result_type.name,
    #                                     "",
    #                                 )
    #                                 data.append(safe_excel_value(raw_value))
    #                             sheet.append(data)

    #             # ========================================================
    #             # EL Image
    #             # ========================================================
    #             elif "EL Image" in procedure_def.name:
    #                 sheet.title = "EL Image"
    #                 sheet.append([
    #                     "Serial Number", "TSD", "LEG", "final_result",
    #                     "Measurement Values"
    #                 ])
    #                 for unit in units:
    #                     tests = normalize_name(
    #                         ProcedureResult.objects.filter(
    #                             unit=unit,
    #                             procedure_definition_id=procedure_def.id,
    #                         )
    #                     )
    #                     if normalized_proc_name:
    #                         tests = tests.filter(normalized_name=normalized_proc_name)
    #                     tests = tests.order_by("test_sequence_definition", "linear_execution_group")

    #                     for test in tests:
    #                         final_result_value = ProcedureResult_FinalResult.objects.filter(
    #                             procedure_result_id=test.id
    #                         ).values_list("final_result", flat=True).first() or "N/A"

    #                         for step_result in test.stepresult_set.all().exclude(archived=True):
    #                             data = [
    #                                 safe_excel_value(unit.serial_number),
    #                                 safe_excel_value(test.test_sequence_definition.name),
    #                                 safe_excel_value(test.name),
    #                                 safe_excel_value(final_result_value)
    #                             ]

    #                             for measurement in step_result.measurementresult_set.all().order_by("report_order"):
    #                                 if measurement.measurement_result_type.name == "result_files":
    #                                     if measurement.result_files.all().count():
    #                                         for azurefile in measurement.result_files.all():
    #                                             if "RAW" in azurefile.file.name:
    #                                                 continue

    #                                             filetype = "DataFiles" if azurefile.file.name.lower().endswith(("xls", "xlsx", "txt", "csv")) else "ImageFiles"

    #                                             name = azurefile.file.name
    #                                             if filetype == "ImageFiles":
    #                                                 temp = azurefile.file.name.split(".")
    #                                                 base_name = ".".join(temp[:-1])
    #                                                 name = "{}-{}-{}.{}".format(
    #                                                     base_name,
    #                                                     test.test_sequence_definition.name,
    #                                                     test.name,
    #                                                     temp[-1]
    #                                                 )

    #                                             bytes_io = BytesIO(azurefile.file.file.read())
    #                                             bytes_io.seek(0)

    #                                             try:
    #                                                 temp_image = Image.open(bytes_io)
    #                                                 if adjust_images:
    #                                                     width, height = temp_image.size
    #                                                     temp_image = temp_image.rotate(-90, expand=True)
    #                                                     temp_image = temp_image.resize((height, width))
    #                                                     temp_image = ImageOps.grayscale(temp_image)
    #                                                     temp_image = ImageOps.autocontrast(temp_image)

    #                                                 image_bytes = BytesIO()
    #                                                 temp_image.save(image_bytes, format="jpeg", quality=75)
    #                                                 temp_image.close()
    #                                                 content = image_bytes.getvalue()
    #                                                 image_bytes.close()
    #                                             except Exception:
    #                                                 bytes_io.seek(0)
    #                                                 content = bytes_io.getvalue()
    #                                             bytes_io.close()

    #                                             files_to_return.append({
    #                                                 "name": f"{work_order.project.number}/{work_order.name}/EL Images/{name}",
    #                                                 "content": content,
    #                                                 "force_zip": True
    #                                             })
    #                                 else:
    #                                     raw_value = getattr(
    #                                         measurement,
    #                                         measurement.measurement_result_type.name,
    #                                         "",
    #                                     )
    #                                     data.append(safe_excel_value(raw_value))
    #                             sheet.append(data)

    #             else:
    #                 return Response(
    #                     {"error": f"Unsupported procedure type for {procedure_def.name}"},
    #                     status=400
    #                 )

    #             # ----------------------------
    #             # Save Excel file (deduplicate)
    #             # ----------------------------
    #             file_stream = BytesIO()
    #             excel_file.save(file_stream)
    #             file_stream.seek(0)

    #             file_name = (
    #                 f"{work_order.name}_{procedure_def.name}.xlsx"
    #                 if not request.data.get("procedures")
    #                 else f"{work_order.name}_{procedure_def.name}_{procedure_name}.xlsx"
    #             )

    #             if not any(f["name"] == file_name for f in files_to_return):
    #                 files_to_return.append({
    #                     "name": file_name,
    #                     "content": file_stream.getvalue(),
    #                     "force_zip": False
    #                 })
    #             file_stream.close()

    #     # ----------------------------
    #     # Return single Excel or ZIP
    #     # ----------------------------
    #     if len(files_to_return) == 1 and not extra_files_exist:
    #         file = files_to_return[0]
    #         response = HttpResponse(
    #             file["content"],
    #             content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    #         )
    #         response["Content-Disposition"] = f'attachment; filename="{file["name"]}"'
    #         return response

    #     mem_zip = BytesIO()
    #     with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
    #         for file in files_to_return:
    #             zf.writestr(file["name"], file["content"])
    #     mem_zip.seek(0)

    #     response = HttpResponse(mem_zip.getvalue(), content_type="application/x-zip-compressed")
    #     response["Content-Disposition"] = (
    #         f'attachment; filename="{work_order.project.number}-{timezone.now().strftime("%b-%d-%Y-%H%M%S")}.zip"'
    #     )
    #     return response

