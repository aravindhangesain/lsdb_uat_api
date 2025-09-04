import json
import os
import zipfile
from io import BytesIO


import numpy as np
import pandas as pd
from PIL import Image, ImageOps
from django.db.models import Q, QuerySet
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from lsdb.models import CrateIntakeImages
from lsdb.models import ModuleProperty 
from lsdb.models import ScannedPannels 
from lsdb.models import UnitType
from lsdb.models import WorkOrder
from lsdb.models import Unit 
from lsdb.models import ProcedureResult
from lsdb.models import ModuleIntakeImages 
from lsdb.models import ModuleIntakeDetails
from lsdb.models import ProcedureResult_FinalResult
from lsdb.utils import ExcelFile
from azure.storage.blob import BlobServiceClient
from lsdb.models import Project


def format_date(date):
    return date.strftime('%Y-%m-%d') if date else 'N/A'

def create_download_file(work_orders: QuerySet[WorkOrder], tsd_ids, unit_ids,
                         procedure_ids, adjust_images=False):
    mem_zip = BytesIO()

    with zipfile.ZipFile(mem_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for work_order in work_orders:
            # Create the Excel files and sheets for various data
            excel_file_units = ExcelFile()
            unit_sheet = excel_file_units.workbook.active
            unit_sheet.title = "Units"
            unit_sheet.append(["Manufacturer", "Technology", "Model",
                               "Serial Number", "Calibration Device Used",
                               "Maximum Voltage[V]", "Width [mm]", "Height [mm]",
                               "Pmax [W]", "VOC [V]", "VMP[V]",
                               "ISC [A]", "IMP [A]","final_result"])

            excel_file_vi = ExcelFile()
            vi_sheet = excel_file_vi.workbook.active
            vi_sheet.title = "Visual Inspection"
            vi_sheet.append(["Serial Number", "TSD", "LEG", "final_result", "Defect", "Category", "Notes", "Images:"])

            excel_file_el = ExcelFile()
            el_sheet = excel_file_el.workbook.active
            el_sheet.title = "EL Image"
            el_sheet.append(
                ["Serial Number", "TSD", "LEG", "final_result","Aperture", "ISO", "Exposure Count", "Injection Current",
                 "Exposure Time"])

            excel_file_wl = ExcelFile()
            wl_sheet = excel_file_wl.workbook.active
            wl_sheet.title = "Wet Leakage"
            wl_sheet.append(["Serial Number", "TSD", "LEG", "final_result","Insulation Resistance", "Passed?", "Test Voltage",
                             "Leakage Current", "Current Trip Setpoint", "Water Temperature"])

            excel_file_dt = ExcelFile()
            dt_sheet = excel_file_dt.workbook.active
            dt_sheet.title = "Diode Test"
            dt_sheet.append(["Serial Number", "TSD", "LEG","final_result", "Forward Voltage", "Reverse Voltage", "Pass?"])

            excel_file_flash = ExcelFile()
            flash_sheet = excel_file_flash.workbook.active
            flash_sheet.title = "Flash"
            flash_sheet.append(
                ["Serial Number", "TSD", "LEG", "final_result","Pmp", "Voc", "Vmp", "Isc", "Imp", "Irradiance", "Temperature"])

            excel_file_color = ExcelFile()
            color_book = excel_file_color.workbook

            if ModuleIntakeDetails.objects.filter(bom=work_order.name,projects=work_order.project_id).exists():
                module_intake = ModuleIntakeDetails.objects.get(bom=work_order.name,projects=work_order.project_id)
                if module_intake:
                    module_intake_images = ModuleIntakeImages.objects.filter(moduleintake_id=module_intake.id)
                    module_intake_images_workbook = Workbook()
                    module_intake_images_sheet = module_intake_images_workbook.active
                    module_intake_images_sheet.title = "ModuleIntakeImages"

                    if module_intake_images.exists():
                        data = []
                        image_streams = {}
                        for module_image in module_intake_images:
                            data.append({
                                'ModuleIntake_id': module_image.moduleintake.id,
                                'Label_name': module_image.label_name,
                                'Image_path': module_image.image_path.name,
                                'Status': module_image.status,
                                'Notes': module_image.notes,
                            })
                            try:
                                blob_service_client = BlobServiceClient.from_connection_string('DefaultEndpointsProtocol=https;AccountName=haveblueazdev;AccountKey=eP954sCH3j2+dbjzXxcAEj6n7vmImhsFvls+7ZU7F4THbQfNC0dULssGdbXdilTpMgaakIvEJv+QxCmz/G4Y+g==;EndpointSuffix=core.windows.net')
                                container_client = blob_service_client.get_container_client('testmedia1')
                                blob_client = container_client.get_blob_client(module_image.image_path.name)
                                image_stream = BytesIO()
                                blob_client.download_blob().readinto(image_stream)
                                image_streams[module_image] = image_stream
                            except Exception as e:
                                print(f"Error downloading image {module_image.image_path.name}: {e}")

                        df = pd.DataFrame(data)
                    else:
                        df = pd.DataFrame(columns=['ModuleIntake_id', 'Label_name', 'Image_path', 'Status', 'Notes'])

                    for row in dataframe_to_rows(df, index=False, header=True):
                        module_intake_images_sheet.append(row)

                    module_intake_images_stream = BytesIO()
                    module_intake_images_workbook.save(module_intake_images_stream)
                    module_intake_images_stream.seek(0)
                    zf.writestr(f'module_intake_images.xlsx', module_intake_images_stream.read())
                    module_intake_images_stream.close()

                    if module_intake_images.exists():
                        for module_image, image_stream in image_streams.items():
                            raw_file_name = module_image.image_path.name.split("/")[-1]
                            base_name, extension = os.path.splitext(raw_file_name)
                            file_name = base_name.split("?")[0].split("&")[0]
                            file_name = file_name.replace("&", "_").replace("?", "_").replace("=", "_")
                            file_name_with_extension = file_name + extension
                            folder_path = f"module_intake_images/{file_name_with_extension}"
                            image_stream.seek(0)
                            zf.writestr(folder_path, image_stream.read())
                            image_stream.close()

                    crate_intake_images = CrateIntakeImages.objects.filter(newcrateintake_id=module_intake.newcrateintake_id)
                    crate_intake_images_workbook = Workbook()
                    crate_intake_images_sheet = crate_intake_images_workbook.active
                    crate_intake_images_sheet.title = "CrateIntakeImages"

                    if crate_intake_images.exists():
                        data = []
                        image_streams = {}
                        for crate_image in crate_intake_images:
                            crate_intake_id = crate_image.newcrateintake.id if crate_image.newcrateintake else None
                            data.append({
                                'CrateIntake_id': crate_intake_id,
                                'Label_name': crate_image.label_name,
                                'Image_path': crate_image.image_path.name if crate_image.image_path else None,
                                'Uploaded_date': format_date(crate_image.uploaded_date),
                                'Project_id': crate_image.project,
                                'Status': crate_image.status,
                                'Notes': crate_image.notes,
                            })
                            try:
                                blob_service_client = BlobServiceClient.from_connection_string('DefaultEndpointsProtocol=https;AccountName=haveblueazdev;AccountKey=eP954sCH3j2+dbjzXxcAEj6n7vmImhsFvls+7ZU7F4THbQfNC0dULssGdbXdilTpMgaakIvEJv+QxCmz/G4Y+g==;EndpointSuffix=core.windows.net')
                                container_client = blob_service_client.get_container_client('testmedia1')
                                blob_client = container_client.get_blob_client(crate_image.image_path.name)
                                image_stream = BytesIO()
                                blob_client.download_blob().readinto(image_stream)
                                image_streams[crate_image] = image_stream
                            except Exception as e:
                                print(f"Error downloading image {crate_image.image_path.name}: {e}")

                        df = pd.DataFrame(data)
                    else:
                        df = pd.DataFrame(columns=['CrateIntake_id', 'Label_name', 'Image_path', 'Uploaded_date', 'Project_id', 'Status', 'Notes'])

                    for row in dataframe_to_rows(df, index=False, header=True):
                        crate_intake_images_sheet.append(row)

                    crate_intake_images_stream = BytesIO()
                    crate_intake_images_workbook.save(crate_intake_images_stream)
                    crate_intake_images_stream.seek(0)
                    zf.writestr(f'crate_intake_images.xlsx', crate_intake_images_stream.read())

                    if crate_intake_images.exists():
                        for crate_image, image_stream in image_streams.items():
                            raw_file_name = crate_image.image_path.name.split("/")[-1]
                            base_name, extension = os.path.splitext(raw_file_name)
                            file_name = base_name.split("?")[0].split("&")[0]
                            file_name = file_name.replace("&", "_").replace("?", "_").replace("=", "_")
                            file_name_with_extension = file_name + extension
                            folder_path = f"crate_intake_images/{file_name_with_extension}"
                            image_stream.seek(0)
                            zf.writestr(folder_path, image_stream.read())
                            image_stream.close()


                    scanned_panels_data = ScannedPannels.objects.filter(module_intake_id=module_intake.id)

                    if scanned_panels_data.exists():
                        scanned_panels_workbook = Workbook()
                        scanned_panels_sheet = scanned_panels_workbook.active
                        scanned_panels_sheet.title = "ScannedPannels"

                        data = []
                        for scanned_panel in scanned_panels_data:
                            data.append({
                                'Moduleintake_id':scanned_panel.module_intake.id if scanned_panel.module_intake else None,
                                'Serial Number': scanned_panel.serial_number,
                                'Status': scanned_panel.status,
                                'TestSequence_id': scanned_panel.test_sequence.id if scanned_panel.test_sequence else None,
                                'Arrival Date':format_date(scanned_panel.arrival_date),
                                'Project Closeout Date':format_date(scanned_panel.project_closeout_date),
                                'EOL Disposition':scanned_panel.eol_disposition
                            })

                        df = pd.DataFrame(data)

                        for row in dataframe_to_rows(df, index=False, header=True):
                            scanned_panels_sheet.append(row)

                        scanned_panels_stream = BytesIO()
                        scanned_panels_workbook.save(scanned_panels_stream)
                        scanned_panels_stream.seek(0)

                        zf.writestr(f'scanned_panels.xlsx', scanned_panels_stream.read())

                    module_type = module_intake.module_type

                    unit_type = UnitType.objects.filter(model=module_type).first()

                    if unit_type:
                        module_property_id = unit_type.module_property_id

                        module_property = ModuleProperty.objects.filter(id=module_property_id).first()

                        if module_property:
                            module_property_workbook = Workbook()
                            module_property_sheet = module_property_workbook.active
                            module_property_sheet.title = "ModuleProperty"

                            data = [{
                                'ModuleProperty_id': module_property.id,
                                'Number of Cells': module_property.number_of_cells,
                                'Nameplate_pmax':module_property.nameplate_pmax,
                                'Module_width':module_property.module_width,
                                'Module_height':module_property.module_height,
                                'System_Voltage':module_property.system_voltage,
                                'Auditor':module_property.auditor,
                                'Audit_date':format_date(module_property.audit_date),
                                'Isc': module_property.isc,
                                'Voc':module_property.voc,
                                'Imp':module_property.imp,
                                'Vmp':module_property.vmp,
                                'Alpha_Isc': module_property.alpha_isc,
                                'Beta_Voc':module_property.beta_voc,
                                'Gamma_pmp':module_property.gamma_pmp,
                                'Cells in Series':module_property.cells_in_series,
                                'Cells in parallel':module_property.cells_in_parallel,
                                'Cell Area':module_property.cell_area,
                                'Bifacial':module_property.bifacial,
                                'Module Technology ID': module_property.module_technology_id,
                            }]

                            df = pd.DataFrame(data)

                            for row in dataframe_to_rows(df, index=False, header=True):
                                module_property_sheet.append(row)

                            module_property_sheet_stream = BytesIO()
                            module_property_workbook.save(module_property_sheet_stream)
                            module_property_sheet_stream.seek(0)

                            zf.writestr(f'module_property.xlsx', module_property_sheet_stream.read())
            else:
                df = pd.DataFrame(columns=['ModuleIntake_id', 'Label_name', 'Image_path', 'Status', 'Notes'])
                df = pd.DataFrame(columns=['CrateIntake_id', 'Label_name', 'Image_path', 'Uploaded_date', 'Project_id', 'Status', 'Notes'])
                df = pd.DataFrame(columns=['ModuleProperty_id', 'Number of Cells', 'Nameplate_pmax', 'Module_width','Module_height', 'System_Voltage', 'Auditor', 'Audit_date', 'Isc', 'Voc', 
                                            'Imp', 'Vmp', 'Alpha_Isc', 'Beta_Voc', 'Gamma_pmp', 'Cells in Series', 
                                            'Cells in parallel', 'Cell Area', 'Bifacial', 'Module Technology ID'])
                df=pd.DataFrame(columns = ['Moduleintake_id', 'Serial Number', 'Status', 'TestSequence_id', 
                                            'Arrival Date', 'Project Closeout Date', 'EOL Disposition'])


            if tsd_ids:
                units = Unit.objects.filter(Q(procedureresult__work_order=work_order) & Q(
                    procedureresult__test_sequence_definition__id__in=tsd_ids)).distinct()
            elif unit_ids:
                units = Unit.objects.filter(
                    Q(procedureresult__work_order=work_order) & Q(id__in=unit_ids)).distinct()
            else:
                units = Unit.objects.filter(procedureresult__work_order=work_order).distinct()

            for unit in units:
                unit_sheet.append([
                    unit.unit_type.manufacturer.name,
                    # unit.unit_type.module_property.module_technology,
                    unit.unit_type.model,
                    unit.serial_number,
                    None,
                    unit.unit_type.module_property.system_voltage,
                    unit.unit_type.module_property.module_width,
                    unit.unit_type.module_property.module_height,
                    unit.unit_type.module_property.nameplate_pmax,
                    unit.unit_type.module_property.voc,
                    unit.unit_type.module_property.vmp,
                    unit.unit_type.module_property.isc,
                    unit.unit_type.module_property.imp
                ])

                if procedure_ids:
                    procedure_query = ProcedureResult.objects.filter(
                        Q(unit=unit) & Q(procedure_definition__id__in=procedure_ids)).order_by(
                        'test_sequence_definition', 'linear_execution_group').select_related(
                        'unit').prefetch_related('stepresult_set__measurementresult_set').distinct()
                else:
                    procedure_query = ProcedureResult.objects.filter(unit=unit).order_by(
                        'test_sequence_definition', 'linear_execution_group').select_related(
                        'unit').prefetch_related('stepresult_set__measurementresult_set')

                for test in procedure_query:
                    final_result = ProcedureResult_FinalResult.objects.filter(procedure_result_id=test.id).values_list('final_result', flat=True).first()
                    final_result_value = final_result if final_result else 'N/A'
                    if "I-V" in test.procedure_definition.name:
                        for step_result in test.stepresult_set.all().exclude(archived=True):
                            data = [unit.serial_number, test.test_sequence_definition.name, test.name, final_result_value]
                            for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                if measurement.measurement_result_type.name == 'result_files':
                                    if measurement.result_files.all().count():
                                        has_result = step_result.measurementresult_set.filter(
                                            result_files__name__icontains="Results")
                                        
                                        for azurefile in measurement.result_files.all():
                                            if has_result and "Results" in azurefile.file.name:
                                                path = "Flash Data/"
                                                if "200" in step_result.name:
                                                    path += "200W/"
                                                elif "Rear" in test.procedure_definition.name:
                                                    path += "Rear/"
                                                path += "FLASH/{}/{}".format(
                                                    test.test_sequence_definition.name,
                                                    test.name
                                                )
                                                bytes = azurefile.file.file.read()
                                                zf.writestr('{}/{}/{}/{}/{}'.format(
                                                    work_order.project.number,
                                                    work_order.name,
                                                    "Flash Data",
                                                    path,
                                                    azurefile.file.name,
                                                ),
                                                    bytes)
                                            elif not has_result:
                                                path = "Flash Data/"
                                                if "200" in step_result.name:
                                                    path += "200W/"
                                                elif "Rear" in test.procedure_definition.name:
                                                    path += "Rear/"
                                                path += "FLASH/{}/{}".format(
                                                    test.test_sequence_definition.name,
                                                    test.name
                                                )
                                                bytes = azurefile.file.file.read()
                                                zf.writestr('{}/{}/{}/{}/{}'.format(
                                                    work_order.project.number,
                                                    work_order.name,
                                                    "Flash Data",
                                                    path,
                                                    azurefile.file.name,
                                                ),
                                                    bytes)
                                            else:
                                                continue
                                        
                                elif measurement.measurement_result_type.name == 'result_datetime':
                                    continue
                                else:
                                    data.append(getattr(measurement, measurement.measurement_result_type.name))
                            flash_sheet.append(data)
                    elif "Wet Leakage" in test.procedure_definition.name:
                        for step_result in test.stepresult_set.all().exclude(archived=True):
                            data = [unit.serial_number, test.test_sequence_definition.name, test.name,final_result_value]
                            for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                data.append(getattr(measurement, measurement.measurement_result_type.name))
                            wl_sheet.append(data)
                    elif "Visual Inspection" in test.procedure_definition.name:
                        skip = False
                        for step_result in test.stepresult_set.all().exclude(archived=True):
                            data = [unit.serial_number, test.test_sequence_definition.name, test.name,final_result_value]
                            for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                if step_result.name == "Inspect Module":
                                    if getattr(measurement, measurement.measurement_result_type.name):
                                        skip = True
                                        data.append("No Defects Observed")
                                        vi_sheet.append(data)
                                elif skip:
                                    break
                                else:
                                        if measurement.measurement_result_type.name == 'result_files':
                                            if measurement.result_files.all().count():
                                                for azurefile in measurement.result_files.all():
                                                    bytes = azurefile.file.file.read()
                                                    zf.writestr('{}/{}/{}/{}'.format(
                                                        work_order.project.number,
                                                        work_order.name,
                                                        "VI Images",
                                                        azurefile.file.name,
                                                    ),
                                                        bytes)
                                                    data.append(azurefile.file.name)
                                    
                                        else:
                                            if measurement.result_defect != None and measurement.measurement_result_type.name == "result_defect":
                                                print(test, step_result, measurement)
                                                print(measurement.result_defect)
                                                data.append(measurement.result_defect.short_name)
                                                data.append(measurement.result_defect.category)
                                            else:
                                                data.append(
                                                    getattr(measurement, measurement.measurement_result_type.name))
                            if skip:
                                break
                            print(data)
                            vi_sheet.append(data)
                    elif "EL Image" in test.procedure_definition.name:
                        for step_result in test.stepresult_set.all().exclude(archived=True):
                            data = [unit.serial_number, test.test_sequence_definition.name, test.name,final_result_value]
                            for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                if measurement.measurement_result_type.name == 'result_files':
                                    if measurement.result_files.all().count():                                    
                                        for azurefile in measurement.result_files.all():
                                            filetype = ''
                                            if "RAW" in azurefile.file.name:
                                                continue

                                            if azurefile.file.name.lower().endswith(
                                                    ('xls', 'xlsx', 'txt', 'csv')):
                                                filetype = 'DataFiles'
                                            else:
                                                filetype = 'ImageFiles'

                                            if filetype == 'ImageFiles':
                                                temp = azurefile.file.name.split(".")
                                                name = temp[0]
                                                for sbstr in range(1, len(temp) - 1):
                                                    name = name + "." + temp[sbstr]

                                                name = "{}-{}-{}.{}".format(name,
                                                                            test.test_sequence_definition.name,
                                                                            test.name, temp[len(temp) - 1])

                                            bytes = BytesIO(azurefile.file.file.read())
                                            bytes.seek(0)

                                            tempImage = Image.open(bytes)
                                            try:
                                                if adjust_images == True:
                                                    width, height = tempImage.size
                                                    tempImage = tempImage.rotate(-90, expand=True)
                                                    tempImage = tempImage.resize((height, width))
                                                    tempImage = ImageOps.grayscale(tempImage)
                                                    tempImage = ImageOps.autocontrast(tempImage)
                                                image_bytes = BytesIO()
                                                tempImage.save(image_bytes, format='jpeg', quality=75)
                                                tempImage.close()
                                                content = image_bytes.getvalue()
                                                image_bytes.close()
                                            except:
                                                bytes.seek(0)
                                                content = bytes.getvalue()
                                            bytes.close()

                                            zf.writestr('{}/{}/{}/{}'.format(
                                                    work_order.project.number,
                                                    work_order.name,
                                                    "EL Images",
                                                    name,
                                                ),
                                                content)
                                      
                                else:
                                    data.append(getattr(measurement, measurement.measurement_result_type.name))
                        el_sheet.append(data)
                    elif "Colorimeter" in test.procedure_definition.name:
                        if "{}".format(unit.serial_number) not in color_book.sheetnames:
                            color_sheet = color_book.create_sheet(title="{}".format(unit.serial_number))
                        for step_result in test.stepresult_set.all().exclude(archived=True):
                            data = [unit.serial_number, test.test_sequence_definition.name, test.name,final_result_value]
                            for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                if step_result.name == "Measure Color" and measurement.measurement_result_type.name == "result_string":
                                    if measurement.result_string == None:
                                        pass
                                        # csv[result.name][result.procedure_definition.name]['body'].append('N/A')
                                    else:
                                        # Convert meassurement string to data values
                                        data = json.loads(measurement.result_string)
                                        data = data["values"]

                                        # Header Rows
                                        color_sheet.append([
                                            "Position",
                                            "L*",
                                            "A*",
                                            "B*"
                                        ])

                                        # Append values to sheet, create columns as we go
                                        l_val, a_val, b_val = [], [], []

                                        color_sheet.append(["Serial Number:", unit.serial_number, "TSD:",
                                                            test.test_sequence_definition.name, "LEG:",
                                                            test.name])
                                        for value in data:
                                            row = [value["position"], value["l_value"], value["a_value"],
                                                   value["b_value"]]
                                            l_val.append(value["l_value"])
                                            a_val.append(value["a_value"])
                                            b_val.append(value["b_value"])
                                            color_sheet.append(row)

                                        color_sheet.append(["Average", np.average(l_val), np.average(a_val),
                                                            np.average(b_val)])

                                        color_sheet.append(["STD", np.std(l_val), np.std(a_val), np.std(b_val)])
                                        color_sheet.append(["", "", "", ""])
                            # color_sheet.append(data)
                    elif "Diode Test" in test.procedure_definition.name:
                        for step_result in test.stepresult_set.all().exclude(archived=True):
                            data = [unit.serial_number, test.test_sequence_definition.name, test.name,final_result_value]
                            for measurement in step_result.measurementresult_set.all().order_by('report_order'):
                                data.append(getattr(measurement, measurement.measurement_result_type.name))
                            dt_sheet.append(data)

           
                            
            # Save the Excel files for each sheet to the zip file
            for title, excel_file in [
                ("Units", excel_file_units),
                ("Visual Inspection", excel_file_vi),
                ("EL Image", excel_file_el),
                ("Wet Leakage", excel_file_wl),
                ("Diode Test", excel_file_dt),
                ("Flash", excel_file_flash),
                ("Colorimeter",excel_file_color)
            ]:
                file_stream = BytesIO()
                excel_file.workbook.save(file_stream)
                file_stream.seek(0)
                zf.writestr(f'{title}.xlsx', file_stream.read())
                file_stream.close()

    filename = timezone.now().strftime('%b-%d-%Y-%H%M%S')
    response = HttpResponse(mem_zip.getvalue(), content_type='application/x-zip-compressed')
    response['Content-Disposition'] = f'attachment; filename={filename}.zip'
    return response